import openpyxl
import json
import sys
from datetime import datetime

def parse_stock_card(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    items = []
    
    # Skip header rows (index 0 and 1)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        
        # ["DATE", "PRODUCT NAME", "EXPIRY ", "QUANTITY", "UNIT", "TOTAL", "UNIT ", "RETAIL ", "RECEIPT", "SUPPLIER"]
        product_name = str(row[1]).strip() if row[1] else ''
        if not product_name or product_name == "''":
            # Sometimes name is omitted if same as previous or just empty
            continue
            
        unit = str(row[4]).strip() if row[4] else 'unit'
        
        unit_cost = 0
        try:
            unit_cost = float(row[6]) if row[6] else 0
        except: pass
        
        quantity = 0
        try:
            quantity = int(row[3]) if row[3] else 0
        except: pass
        
        supplier = str(row[9]).strip() if len(row) > 9 and row[9] else ''
        if supplier == "''": supplier = ''
            
        items.append({
            "name": product_name,
            "unit_of_measure": unit,
            "unit_cost": unit_cost,
            "quantity_on_hand": quantity,
            "supplier": supplier,
            "is_billable": True,
            "is_active": True,
            "category": "OTHER"
        })
        
    return items

def parse_pricelist(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    services = []
    
    # Skip header (index 0)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 1: continue
        
        name = str(row[0]).strip() if row[0] else ''
        if not name: continue
            
        price = 0
        try:
            price = float(row[1]) if len(row) > 1 and row[1] else 0
        except: pass
            
        services.append({
            "service_name": name,
            "price": price,
            "category": "PHARMACY", # Assuming most items in pricelist are pharmacy/consumables
            "price_type": "FIXED",
            "is_active": True
        })
        
    return services

if __name__ == "__main__":
    stock_card_path = sys.argv[1]
    pricelist_path = sys.argv[2]
    out_path = sys.argv[3]
    
    data = {
        "inventory": parse_stock_card(stock_card_path),
        "services": parse_pricelist(pricelist_path)
    }
    
    with open(out_path, 'w') as f:
        json.dump(data, f, indent=2)
        
    print(f"Parsed {len(data['inventory'])} inventory items and {len(data['services'])} services.")
