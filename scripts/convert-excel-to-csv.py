#!/usr/bin/env python3
"""
Convert Excel file to CSV format for patient import.
Usage: python3 scripts/convert-excel-to-csv.py "NS CLIENT FILES - (3).xlsx" patient-data.csv
"""

import openpyxl
import csv
import sys
import os

def convert_excel_to_csv(excel_path, csv_path):
    """Convert Excel file to CSV format."""
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Open CSV file for writing
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)
            
            # Write all rows from Excel to CSV
            for row in ws.iter_rows(values_only=True):
                # Convert all values to strings and handle None values
                csv_row = [str(cell) if cell is not None else '' for cell in row]
                csv_writer.writerow(csv_row)
        
        print(f"Successfully converted {excel_path} to {csv_path}")
        print(f"Total rows: {ws.max_row}")
        return True
        
    except Exception as e:
        print(f"Error converting Excel to CSV: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 scripts/convert-excel-to-csv.py <excel_file> <csv_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    csv_file = sys.argv[2]
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found")
        sys.exit(1)
    
    success = convert_excel_to_csv(excel_file, csv_file)
    sys.exit(0 if success else 1)
